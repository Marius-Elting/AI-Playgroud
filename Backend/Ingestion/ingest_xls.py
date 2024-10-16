from typing import Any
from openpyxl import Workbook, load_workbook


class ExcelIngestion:

    def extract_tables(self, file)-> list[Any]:
        wb: Workbook = load_workbook(filename=file.file)
        ws = wb["SalesOrders"]
        for ws in wb:
            table_contents: list[Any] = []
            for table_name, data in ws.tables.items():
                data = ws[data]
                content = [[cell.value for cell in ent] for ent in data]
                

                header: list[Any] = [content.replace(" ", "_") for content in content[0]]
                for content_chunk in content:
                    for i in range(len(content_chunk)):
                        content_chunk[i] = str(object=content_chunk[i])
                    
                    print(content_chunk)
                chunks: list[Any] = []
                for i in range(1, len(content), 10):
                    if i == 0:
                        continue
                    current_chunk = content[i:i+10]
                    combinded_chunk = [header] + current_chunk
                    chunks.append(combinded_chunk)
                

                chunk_strings: list[Any] = []
                for chunk in chunks:
                    chunk_strings.append("| "+" | \n | ".join([" | ".join(row) for row in chunk]))

                table_contents.append(chunk_strings)

            return table_contents
        return []